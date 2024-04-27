import pandas as pd
from openpyxl import load_workbook
import os

path = os.getcwd()

class ExcelProcessor:
    def __init__(self):
        # Initialize any attributes you need
        self.all_dfs = []  # List to store DataFrames

    def process_all_excel_files(self, folder_path):
        # Iterate through each file in the folder
        for filename in os.listdir(folder_path):
            if filename.endswith(".xlsx"):
                # Construct the full file path
                file_path = os.path.join(folder_path, filename)

                # Call the function to extract the table from each Excel file
                df = self.extract_table_from_excel(file_path)

                # Check if there's data in the DataFrame before appending
                if not df.empty:
                    # Add a 2-row gap by appending two empty rows
                    empty_rows = pd.DataFrame(index=range(2), columns=df.columns)
                    self.all_dfs.extend([df,empty_rows])

        # Concatenate all DataFrames in the list into a single DataFrame
        combined_df = pd.concat(self.all_dfs, ignore_index=True)
        combined_df = combined_df.fillna(' ')

        # Display or save the combined DataFrame
        print("\nCombined DataFrame:")
        print(combined_df)

        return combined_df
        # combined_df.to_excel("MULTI_TABLE_COMBINED.xlsx", index=False)

    def copy_and_format_data(self,file_path,shipping_comp_name):

        if "SITC" in shipping_comp_name:
            print("()(------------------->>>>>>>)",shipping_comp_name)
            # Read the Excel file into a pandas DataFrame
            xls = pd.ExcelFile(file_path)

            # Find the sheets containing 'INST' in their names
            desired_sheets = [sheet for sheet in xls.sheet_names if 'At' in sheet]

            if not desired_sheets:
                print("No sheet found with 'At' in its name.")
            else:
                self.output_directory = path + f'/business_logic/excel_extracter/Test_Excel_Data_Extr'
                os.makedirs(self.output_directory, exist_ok=True)

                # Iterate over desired sheets
                for sheet_name in desired_sheets:
                    # Read the data from the desired sheet
                    df = pd.read_excel(file_path, sheet_name=sheet_name)

                    # Create a new Excel file in the specified directory
                    new_file_path = os.path.join(self.output_directory, f'output_file.xlsx')
                    with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
                        # Write the data to a new sheet
                        df.to_excel(writer, sheet_name='NewSheet', index=False)

                        # Access the new workbook and sheet to apply formatting
                        workbook = writer.book
                        worksheet = writer.sheets['NewSheet']

                        # Adjust column widths based on the data
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:  # Necessary to avoid error on empty cells
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                    print(f"Data from '{sheet_name}' sheet has been copied to '{new_file_path}' in the 'NewSheet'.")
        else:
            print("Find the inst in the Excel sheet")
            xls = pd.ExcelFile(file_path)

            # Find the sheets containing 'INST' in their names
            desired_sheets = [sheet for sheet in xls.sheet_names if 'INST' in sheet]

            if not desired_sheets:
                print("No sheet found with 'INST' in its name.")
            else:
                self.output_directory = path + f'/business_logic/excel_extracter/Test_Excel_Data_Extr'
                os.makedirs(self.output_directory, exist_ok=True)

                # Iterate over desired sheets
                for sheet_name in desired_sheets:
                    # Read the data from the desired sheet
                    df = pd.read_excel(file_path, sheet_name=sheet_name)

                    # Create a new Excel file in the specified directory
                    new_file_path = os.path.join(self.output_directory, f'output_file.xlsx')
                    with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
                        # Write the data to a new sheet
                        df.to_excel(writer, sheet_name='NewSheet', index=False)

                        # Access the new workbook and sheet to apply formatting
                        workbook = writer.book
                        worksheet = writer.sheets['NewSheet']

                        # Adjust column widths based on the data
                        for column in worksheet.columns:
                            max_length = 0
                            column = [cell for cell in column]
                            for cell in column:
                                try:  # Necessary to avoid error on empty cells
                                    if len(str(cell.value)) > max_length:
                                        max_length = len(cell.value)
                                except:
                                    pass
                            adjusted_width = (max_length + 2)
                            worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                    print(f"Data from '{sheet_name}' sheet has been copied to '{new_file_path}' in the 'NewSheet'.")


    def extract_value_to_right(self, sheet, row, col):
        # Search the first three cells to the right
        for offset in range(1, 4):
            col_num = col + offset
            if col_num <= sheet.max_column:
                value = sheet.cell(row=row, column=col_num).value

                # Check for merged cells
                if value is not None and sheet.merged_cells and sheet.cell(row=row, column=col_num).coordinate in sheet.merged_cells:
                    merged_range = sheet.merged_cells[sheet.cell(row=row, column=col_num).coordinate]
                    # Extract the value from the top-left cell of the merged range
                    value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value

                if value is not None:
                    return str(value)

        # Return an empty string if no non-empty cell is found
        return ""

    def find_and_print_values(self, *target_values):
        # Load the existing workbook (output_file.xlsx)
        output_file_path = self.output_directory + '/output_file.xlsx'
        workbook = load_workbook(output_file_path)
        BL_issue_h31 = []
        final_values = []
        # Iterate through all sheets in the newly created workbook
        for sheet_name in workbook.sheetnames:
            if 'NewSheet' in sheet_name:  # Assuming the newly created sheet is named 'NewSheet'
                sheet = workbook[sheet_name]
                
                for target_value in target_values:
                    # Find the coordinates of the cell with the target value
                    target_value_row = None
                    target_value_col = None
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value == target_value:
                                target_value_row = cell.row
                                target_value_col = cell.column
                                break

                    # If the target value is found, extract the first three values to the right
                    if target_value_row is not None and target_value_col is not None:
                        value_to_right = self.extract_value_to_right(sheet, target_value_row, target_value_col)
                        
                        BL_issue_h31.append(value_to_right)
                        print(f"Sheet: {sheet_name}, Cell coordinates of '{target_value}': ({target_value_row}, {target_value_col})")
                        print(f"Value to the right: {value_to_right}")

                        # Check for values under 'CONSIGNEE :' and 'NOTIFY PARTY :'
                        if target_value == 'CONSIGNEE :' or target_value == 'NOTIFY PARTY :':
                            value_under_target = self.extract_value_to_right(sheet, target_value_row + 1, target_value_col)
                            value_under_target1 = self.extract_value_to_right(sheet, target_value_row + 2, target_value_col)

                            if value_under_target:
                                # Concatenate the values if there are values under 'CONSIGNEE :' or 'NOTIFY PARTY :'
                                final_value = f"{value_to_right} {value_under_target} {value_under_target1}"
                                final_values.append(final_value)
                                print(f"Final concatenated value: {final_value}")
                            else:
                                final_value = f"{value_to_right}"
                                final_values.append(final_value)
                                print("No value under the target cell.")
        print(f"Final concatenated values------->>>: {final_values}")
        # Don't forget to close the workbook when you're done with it
        workbook.close()
        return BL_issue_h31, final_values


    def extract_table_from_excel(self, file_path):
        # Find the sheet containing 'INST' in its name
        xls = pd.ExcelFile(file_path)
        desired_sheet = [sheet for sheet in xls.sheet_names if 'INVOICE' in sheet or 'At' in sheet]
        xls.close()

        if not desired_sheet:
            print(f"No sheet found with 'INVOICE' in its name in file: {file_path}")
            return pd.DataFrame()  #Return an empty DataFrame
        
        elif "At" in desired_sheet[0]:
            # Specify the range of cells in the table
            start_row = 3  # Replace with the actual starting row of your table
            end_row = 13   # Replace with the actual ending row of your table
            start_col = 'A' # Replace with the actual starting column of your table
            end_col = 'P'   # Replace with the actual ending column of your table

            # Read the data from the desired sheet and specified range
            df = pd.read_excel(file_path, sheet_name=desired_sheet[0], skiprows=start_row-1, usecols=f"{start_col}:{end_col}", nrows=end_row-start_row+1)

            # Find the row index where 'Total' is present in any cell
            total_row_index = df.apply(lambda row: row.astype(str).str.contains('UNITS', case=False).any(), axis=1).idxmax()

            # Exclude the row with 'Total' if found
            if total_row_index is not None:
                df = df.loc[:total_row_index - 4]

            # df = df.fillna('')
            # Display the DataFrame or perform further operations as needed
            print("DataFrame000000>>>>>>",df)
            return df

        else:
            # Specify the range of cells in the table
            start_row = 13  # Replace with the actual starting row of your table
            end_row = 24   # Replace with the actual ending row of your table
            start_col = 'A' # Replace with the actual starting column of your table
            end_col = 'Q'   # Replace with the actual ending column of your table

            # Read the data from the desired sheet and specified range
            df = pd.read_excel(file_path, sheet_name=desired_sheet[0], skiprows=start_row-1, usecols=f"{start_col}:{end_col}", nrows=end_row-start_row+1)

            # Find the row index where 'Total' is present in any cell
            total_row_index = df.apply(lambda row: row.astype(str).str.contains('Total', case=False).any(), axis=1).idxmax()

            # Exclude the row with 'Total' if found
            if total_row_index is not None:
                df = df.loc[:total_row_index - 1]

            df = df.fillna('')
            # Display the DataFrame or perform further operations as needed
            print("DataFrame000000>>>>>>",df)

            return df
    
