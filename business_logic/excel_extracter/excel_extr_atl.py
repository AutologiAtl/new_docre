import logging
import traceback
import pandas as pd
from openpyxl import load_workbook
import os

path = os.getcwd()
sep = os.path.sep

class ExcelProcessor:
    def __init__(self):
        # Initialize any attributes you need
        self.all_dfs = []  # List to store DataFrames

    def process_all_excel_files(self, folder_path):
        print("pro",folder_path)
        # Iterate through each file in the folder
        for filename in os.listdir(folder_path):
            print("pro1",filename)
            if filename.endswith((".xlsx")):
                # Construct the full file path
                file_path = os.path.join(folder_path, filename)

                # Call the function to extract the table from each Excel file
                df = self.extract_table_from_excel(file_path)

                # Check if there's data in the DataFrame before appending
                if not df.empty:
                    # Add a 2-row gap by appending two empty rows
                    empty_rows = pd.DataFrame(index=range(2), columns=df.columns)
                    self.all_dfs.extend([df,empty_rows])

        #Concatenate all DataFrames in the list into a single DataFrame
        combined_df = pd.concat(self.all_dfs, ignore_index=True)
        print("combined_df",combined_df)
        combined_df = combined_df.fillna(' ')

        # Display or save the combined DataFrame
        print("\nCombined DataFrame:")
        print(combined_df)


        return combined_df
        # combined_df.to_excel("MULTI_TABLE_COMBINED.xlsx", index=False)
    def copy_and_format_data(self, file_path):
        
        self.file_path = file_path
        xls = pd.ExcelFile(self.file_path)
        sheet_names = xls.sheet_names

        if 'Si' in sheet_names or 'Invoice' in sheet_names or 'INST-1' in sheet_names:
            condition = lambda sheet: sheet == 'Si' or sheet == 'Invoice' or sheet == 'INST-1'

        # elif 'MONG2359645' in sheet_names or 'GEO2361810' in sheet_names or 'THAIL2353544' in sheet_names:
        #     condition = lambda sheet: sheet == 'MONG2359645' or sheet == 'GEO2361810' or sheet == 'THAIL2353544'

        else:
            condition = lambda sheet: 'INST-3' in sheet or sheet == 'INVOICE' or sheet == 'MONG2359645'

        desired_sheets = [sheet for sheet in xls.sheet_names if condition(sheet)]

        if not desired_sheets:
            print(f"No sheet found with '{'Si' if 'SITC' in self.file_path else 'INST-1'}' in its name.")
        else:
            self.process_sheets(desired_sheets)

    def process_sheets(self, desired_sheets):

        file_path = self.file_path
        self.output_directory = os.path.join(path+ f'{sep}business_logic{sep}excel_extracter{sep}Test_Excel_Data_Extr')
        os.makedirs(self.output_directory, exist_ok=True)

        for sheet_name in desired_sheets:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            new_file_path = os.path.join(self.output_directory, f'output_file.xlsx')

            with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='NewSheet', index=False)

                workbook = writer.book
                worksheet = writer.sheets['NewSheet']

                for column in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in column)
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                print(f"Data from '{sheet_name}' sheet has been copied to '{new_file_path}' in the 'NewSheet'.")


    def extract_value_to_right(self, sheet, row, col):
        # Search the first three cells to the right
        for offset in range(1, 4):
            col_num = col + offset
            if col_num <= sheet.max_column:
                value = sheet.cell(row=row, column=col_num).value

                # Check for merged cells
                if value is not None and sheet.merged_cells and sheet.cell(row=row, column=col_num).coordinate in sheet.merged_cells:
                    merged_range = sheet.merged_cells[sheet.cell(row=row, column=col_num).coordinate]
                    # Extract the value from the top-left cell of the merged range
                    value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value

                if value is not None:
                    return str(value)

        # Return an empty string if no non-empty cell is found
        return ""

    def find_and_print_values(self, *target_values):
        try:
            # Load the existing workbook (output_file.xlsx)
            output_file_path = self.output_directory + f'{sep}output_file.xlsx'
            workbook = load_workbook(output_file_path)
            final_values = []
            # Iterate through all sheets in the newly created workbook
            for sheet_name in workbook.sheetnames:
                if 'NewSheet' in sheet_name:  # Assuming the newly created sheet is named 'NewSheet'
                    sheet = workbook[sheet_name]
                    
                    for target_value in target_values:
                        print(f"Searching for: {target_value}")
                        target_value_row = None
                        target_value_col = None
                    
                        for row in sheet.iter_rows():
                            for cell in row:
                                # Convert cell value and target value to lowercase for case-insensitive comparison
                                if str(cell.value).lower() == str(target_value).lower():
                                    target_value_row = cell.row
                                    target_value_col = cell.column
                                    break

                        try:
                            if target_value_row is not None and target_value_col is not None:
                                value_to_right = self.extract_value_to_right(sheet, target_value_row, target_value_col)

                                print(f"Value to the right: {value_to_right}")

                                if target_value == 'FREIGHT:  ' or target_value == 'Vessel : ':
                                    value_under_target0 = self.extract_value_to_right(sheet, target_value_row, target_value_col+2)
                                    # value_under_target1 = self.extract_value_to_right(sheet, target_value_row + 2, target_value_col)

                                    if value_under_target0:
                                        # Concatenate the values if there are values under 'CONSIGNEE :' or 'NOTIFY PARTY :'
                                        final_FREIGHT_Vessel_values = f"{value_to_right} {value_under_target0}"
                                        final_values.append(final_FREIGHT_Vessel_values)
                                    else:
                                        final_value = f"{value_to_right}"
                                        final_values.append(final_value)
                                        print("No value under the target cell.")
                                        
                                elif target_value == 'Shipper : ':
                                    value_under_target2 = self.extract_value_to_right(sheet, target_value_row + 1, target_value_col)
                                    value_under_target3 = self.extract_value_to_right(sheet, target_value_row + 2, target_value_col)
                                    value_under_target4 = self.extract_value_to_right(sheet, target_value_row + 3, target_value_col)
                                    value_under_target5 = self.extract_value_to_right(sheet, target_value_row + 4, target_value_col)
                                    value_under_target6= self.extract_value_to_right(sheet, target_value_row + 5, target_value_col)

                                    if value_under_target2:
                                        # Concatenate the values if there are values under 'CONSIGNEE :' or 'NOTIFY PARTY :'
                                        final_Shipper_value = f"{value_to_right} {value_under_target2} {value_under_target3} {value_under_target4} {value_under_target5} {value_under_target6}"
                                        final_values.append(final_Shipper_value)
                                    else:
                                        final_value = f"{value_to_right}"
                                        final_values.append(final_value)
                                        print("No value under the target cell.")

                                # Check for values under 'CONSIGNEE :' and 'NOTIFY PARTY :'
                                if (target_value == 'Consignee : ' or target_value == 'Notify : ') or (target_value == 'CONSIGNEE :' or target_value == 'NOTIFY PARTY :'):
                                    print(f"target_value:{target_value}")
                                    value_under_target0 = self.extract_value_to_right(sheet, target_value_row + 1, target_value_col)
                                    value_under_target1 = self.extract_value_to_right(sheet, target_value_row + 2, target_value_col)

                                    if value_under_target0:
                                        # Concatenate the values if there are values under 'CONSIGNEE :' or 'NOTIFY PARTY :'
                                        final_value = f"{value_to_right} {value_under_target0} {value_under_target1}"
                                        print(f"final_value:{final_value}")
                                        final_values.append(final_value)
                                    else:
                                        final_value = f"{value_to_right}"
                                        final_values.append(final_value)
                                        print("No value under the target cell.")

                                elif target_value == 'B/L ISSUE BY :':
                                    BL_issue_Value = self.extract_value_to_right(sheet, target_value_row, target_value_col)
                                    if BL_issue_Value:
                                        # Concatenate the values if there are values under 'CONSIGNEE :' or 'NOTIFY PARTY :'
                                        BL_issue_h31_1 = f"{BL_issue_Value}"
                                        final_values.append(BL_issue_h31_1)
                                    else:
                                        final_values.append("Not Found!")
                        except:
                            print(f"The block in the exception ------->\n{target_value}")
                            # Check for values under 'CONSIGNEE :' and 'NOTIFY PARTY :'
                            if target_value == 'CONSIGNEE :' or target_value == 'NOTIFY PARTY :':
                                print(f"target_value:{target_value}")
                                value_under_target0 = self.extract_value_to_right(sheet, target_value_row + 1, target_value_col)
                                value_under_target1 = self.extract_value_to_right(sheet, target_value_row + 2, target_value_col)

                                if value_under_target0:
                                    # Concatenate the values if there are values under 'CONSIGNEE :' or 'NOTIFY PARTY :'
                                    final_value = f"{value_to_right} {value_under_target0} {value_under_target1}"
                                    print(f"final_value:{final_value}")
                                    final_values.append(final_value)
                                else:
                                    final_value = f"{value_to_right}"
                                    final_values.append(final_value)
                                    print("No value under the target cell.")

                            elif target_value == 'B/L :' or target_value == 'B/L ISSUE BY :':
                                BL_issue_Value = self.extract_value_to_right(sheet, target_value_row, target_value_col)
                                if BL_issue_Value:
                                    # Concatenate the values if there are values under 'CONSIGNEE :' or 'NOTIFY PARTY :'
                                    BL_issue_h31_1 = f"{BL_issue_Value}"
                                    final_values.append(BL_issue_h31_1)
                                else:
                                    final_values.append("Not Found!")    
                            
            # Printing the DataFrame
            print(f"concatenated final_values------->>>: \n{final_values}")
            # Don't forget to close the workbook when you're done with it
            workbook.close()
            return final_values
        except Exception as e:
            print(f"Error Exception in excel_extr_atl file: {e}")
            logging.basicConfig(filename='excel_extr_atl.log', level=logging.DEBUG, format='%(asctime)s - %(levelname)s - %(message)s')
            logging.error(f'Here is the Error \n{traceback.print_exc()} \n\n Exception \n{e}')
            print("Created logfile...")
        return None, None


    def extract_table_from_excel(self, file_path):
        # Check file extension to determine the engine
        if file_path.lower().endswith('.xlsx'):
            engine = 'openpyxl'
        else:
            print(f"Unsupported file format: {file_path}")
            return pd.DataFrame()

        xls = pd.ExcelFile(file_path,engine=engine)
        # desired_sheet = [sheet for sheet in xls.sheet_names if 'INVOICE' in sheet or 'At' in sheet or 'AT' in sheet or 'MONG2359645' in sheet]
        desired_sheet = [sheet for sheet in xls.sheet_names if 'INVOICE' in sheet or 'At' in sheet or 'AT' in sheet or 'MONG' in sheet]
        xls.close()

        try:
            # if not desired_sheet:
            #     print(f"No sheet found with 'INVOICE' in its name in file: {file_path}")
            #     return pd.DataFrame()  #Return an empty DataFrame
            
            if "At" in desired_sheet[0] or "AT" in desired_sheet[0]:
                df = pd.read_excel(file_path, sheet_name=desired_sheet[0])                
                # Specify the range of cells in the table
                start_row = 3
                # Apply the search using regular expressions
                end_row = df.apply(lambda row: row.isin(['JPY', 'KGS', 'UNITS']).any(), axis=1).idxmax()
                # end_row = df.apply(lambda row: row.astype(str).str.contains('"TOTAL : " 0 "UNIT(S)"', case=False).any(), axis=1).idxmax()
                print(f"end_row   \n{end_row}")
                end_col = chr(ord('A') + df.shape[1] - 1)

                df = pd.read_excel(file_path, sheet_name=desired_sheet[0], skiprows=start_row-1, usecols=f"A:{end_col}", nrows=end_row-start_row+1)
                print("DATAFRAME1112223334455566--->",df)
                # Exclude the row with 'Total' if found
                if end_row is not None:
                    df = df.loc[:end_row - 4]

                print(f"DataFrame for <{desired_sheet[0]}> :>>>>>>", df)
                return df
                
            

           
            

            else:
                df = pd.read_excel(file_path, sheet_name=desired_sheet[0])
                start_row = 13
                end_row = df.apply(lambda row: row.isin(['TOTAL', 'JPY', 'KGS']).any(), axis=1).idxmax()
                end_col = chr(ord('A') + df.shape[1] - 1)

                df = pd.read_excel(file_path, sheet_name=desired_sheet[0], skiprows=start_row-1, usecols=f"A:{end_col}", nrows=end_row-start_row+1)

                # Convert end_col from letter to column index
                end_col_index = ord(end_col) - ord('A') + 1

                # Slice the DataFrame using numerical column indices
                df = df.iloc[:, :end_col_index]

                print("DataFrame for <{}>: {}".format(desired_sheet[0], df))

                return df
        except:
            if len(xls.sheet_names) > 1:
                second_sheet_name = xls.sheet_names[0]  # Sheets are 0-indexed, so the second sheet is at index 1
                print("The second sheet's name is:", second_sheet_name)
                # Specify the range of cells in the table
                df = pd.read_excel(file_path, sheet_name=second_sheet_name)
                print("check df df",df)
                start_row = 13
                print("start row",start_row)
                # Apply the search using regular expressions
                end_row = df.apply(lambda row: row.isin(['DESCRIPTION:', 'SHIPPING FROM:']).any(), axis=1).idxmax()
                # end_row = df.apply(lambda row: row.astype(str).str.contains('"TOTAL : " 0 "UNIT(S)"', case=False).any(), axis=1).idxmax()
                print(f"end_row   \n{end_row}")
                end_col = chr(ord('A') + df.shape[1] - 1)

                df = pd.read_excel(file_path, sheet_name=second_sheet_name, skiprows=start_row-1, usecols=f"C:{end_col}", nrows=end_row-start_row+1)
                print("DATAFRAME1112223334455566--->",df)
                # Exclude the row with 'Total' if found
                if end_row is not None:
                    df = df.loc[:end_row - 4]

                print(f"DataFrame for <{second_sheet_name}> :>>>>>>", df)
                return df
            
    
