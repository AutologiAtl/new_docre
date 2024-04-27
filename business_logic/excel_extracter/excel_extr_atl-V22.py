import traceback
import pandas as pd
from openpyxl import load_workbook
import os

path = os.getcwd()
sep = os.path.sep

class ExcelProcessor:
    def __init__(self):
        # Initialize any attributes you need
        self.all_dfs = []  # List to store DataFrames

    def process_all_excel_files(self, folder_path):
        # Iterate through each file in the folder
        for filename in os.listdir(folder_path):
            if filename.endswith((".xlsx",".xls")):
                # Construct the full file path
                file_path = os.path.join(folder_path, filename)

                # Call the function to extract the table from each Excel file
                df = self.extract_table_from_excel(file_path)

                # Check if there's data in the DataFrame before appending
                if not df.empty:
                    # Add a 2-row gap by appending two empty rows
                    empty_rows = pd.DataFrame(index=range(2), columns=df.columns)
                    self.all_dfs.extend([df,empty_rows])

        # Concatenate all DataFrames in the list into a single DataFrame
        combined_df = pd.concat(self.all_dfs, ignore_index=True)
        combined_df = combined_df.fillna(' ')

        # Display or save the combined DataFrame
        print("\nCombined DataFrame:")
        print(combined_df)

        return combined_df
        # combined_df.to_excel("MULTI_TABLE_COMBINED.xlsx", index=False)
    def copy_and_format_data(self, file_path, shipping_comp_name=None):
        print("()(------------------->>>>>>>)", shipping_comp_name)
        self.file_path = file_path

        xls = pd.ExcelFile(file_path)
        shipping_comp_name = "MSC"

        if shipping_comp_name:
            condition = lambda sheet: sheet == 'INVOICE'
        elif shipping_comp_name:
            condition = lambda sheet: sheet == 'AT'
        elif shipping_comp_name:
            condition = lambda sheet: sheet == 'At'
        elif shipping_comp_name:
            condition = lambda sheet: sheet == 'Si'
        else:
            condition = lambda sheet: 'INST-1' in sheet

        desired_sheets = [sheet for sheet in xls.sheet_names if condition(sheet)]

        if not desired_sheets:
            print(f"No sheet found with '{'Si' if 'SITC' in shipping_comp_name else 'INST-1'}' in its name.")
        else:
            self.process_sheets(desired_sheets)

    def process_sheets(self, desired_sheets):

        file_path = self.file_path
        self.output_directory = os.path.join(path+ '/business_logic/excel_extracter/Test_Excel_Data_Extr')
        os.makedirs(self.output_directory, exist_ok=True)

        for sheet_name in desired_sheets:
            df = pd.read_excel(file_path, sheet_name=sheet_name)
            new_file_path = os.path.join(self.output_directory, f'output_file.xlsx')

            with pd.ExcelWriter(new_file_path, engine='openpyxl') as writer:
                df.to_excel(writer, sheet_name='NewSheet', index=False)

                workbook = writer.book
                worksheet = writer.sheets['NewSheet']

                for column in worksheet.columns:
                    max_length = max(len(str(cell.value)) for cell in column)
                    adjusted_width = max_length + 2
                    worksheet.column_dimensions[column[0].column_letter].width = adjusted_width

                print(f"Data from '{sheet_name}' sheet has been copied to '{new_file_path}' in the 'NewSheet'.")


    def extract_value_to_right(self, sheet, row, col):
        # Search the first three cells to the right
        for offset in range(1, 4):
            col_num = col + offset
            if col_num <= sheet.max_column:
                value = sheet.cell(row=row, column=col_num).value

                # Check for merged cells
                if value is not None and sheet.merged_cells and sheet.cell(row=row, column=col_num).coordinate in sheet.merged_cells:
                    merged_range = sheet.merged_cells[sheet.cell(row=row, column=col_num).coordinate]
                    # Extract the value from the top-left cell of the merged range
                    value = sheet.cell(row=merged_range.min_row, column=merged_range.min_col).value

                if value is not None:
                    return str(value)

        # Return an empty string if no non-empty cell is found
        return ""

    def find_and_print_values(self, *target_values):
        # Load the existing workbook (output_file.xlsx)
        output_file_path = self.output_directory + f'{sep}output_file.xlsx'
        workbook = load_workbook(output_file_path)

        
        BL_issue_h31 = []
        final_values = []
        # Iterate through all sheets in the newly created workbook
        for sheet_name in workbook.sheetnames:
            if 'NewSheet' in sheet_name:  # Assuming the newly created sheet is named 'NewSheet'
                sheet = workbook[sheet_name]
                extracted_cells = []
                for target_value in target_values:
                    # Find the coordinates of the cell with the target value
                    target_value_row = None
                    target_value_col = None
                    target_cell = None
                    for row in sheet.iter_rows():
                        for cell in row:
                            if cell.value == target_value:
                                target_value_row = cell.row
                                target_value_col = cell.column
                                break           
                    
                    # If the target value is found, extract the first three values to the right
                    if target_value_row is not None and target_value_col is not None:
                        value_to_right = self.extract_value_to_right(sheet, target_value_row, target_value_col)
                        
                        BL_issue_h31.append(value_to_right)
                        print(f"Sheet: {sheet_name}, Cell coordinates of '{target_value}': ({target_value_row}, {target_value_col})")
                        print(f"Value to the right: {value_to_right}")

                        # Check for values under 'CONSIGNEE :' and 'NOTIFY PARTY :'
                        if target_value == 'CONSIGNEE :' or target_value == 'NOTIFY PARTY :':
                            value_under_target = self.extract_value_to_right(sheet, target_value_row + 1, target_value_col)
                            value_under_target1 = self.extract_value_to_right(sheet, target_value_row + 2, target_value_col)

                            if value_under_target:
                                # Concatenate the values if there are values under 'CONSIGNEE :' or 'NOTIFY PARTY :'
                                final_value = f"{value_to_right} {value_under_target} {value_under_target1}"
                                final_values.append(final_value)
                                print(f"Final concatenated value: {final_value}")
                            else:
                                final_value = f"{value_to_right}"
                                final_values.append(final_value)
                                print("No value under the target cell.")

                # Search for the target cell containing "MARKS & NO.S"
                target_cell = None
                for row in sheet.iter_rows():
                    # vessel_number = row('C21').value
                    for cell in row:
                        if cell.value == "YUGENGAISYA MATSUDA TRADING":
                            target_cell = cell
                            break
                    if target_cell:
                        break
                
                if not target_cell:
                    print("Target cell not found")
                
                    return None
                
                # Extract the contents of the four cells below the target cell
                extracted_cells = []
                for i in range(0, 5):  # Extract four cells below the target cell
                    row_index = target_cell.row + i
                    column_index = target_cell.column
                    cell_value = sheet.cell(row=row_index, column=column_index).value
                    extracted_cells.append(cell_value)
                print(f"Final extracted_cells values------->>>: {extracted_cells},'\n'-{target_cell}")
        print(f"Final concatenated values------->>>: {final_values}")
        # Don't forget to close the workbook when you're done with it
        workbook.close()
        return BL_issue_h31, final_values
    
    # -------- MARKS & NUMBERS CODE END HERE --------

    def extract_table_from_excel(self, file_path):
        
        if file_path.lower().endswith('.xlsx'):
            engine = 'openpyxl'
        else:
            print(f"Unsupported file format: {file_path}")
            return pd.DataFrame()
        print("Check file extension to determine the engine",file_path)
        xls = pd.ExcelFile(file_path,engine=engine)
        # print("traceback.print_exc()")    
        desired_sheet = [sheet for sheet in xls.sheet_names if 'INVOICE' in sheet or 'At' in sheet or 'AT' in sheet]
        print("desired_sheet==== desired_sheet== desired_sheet",desired_sheet)
        xls.close()

        try:
            if not desired_sheet:
                print(f"No sheet found with 'INVOICE' in its name in file: {file_path}")
                return pd.DataFrame()  #Return an empty DataFrame
            
            elif "At" in desired_sheet[0] or "AT" in desired_sheet[0]:

                df = pd.read_excel(file_path, sheet_name=desired_sheet[0])

                # Specify the range of cells in the table
                start_row = 3
                end_row = 13

                # Dynamically set the end_col based on the number of columns
                end_col = chr(ord('A') + df.shape[1] - 1)
                df = pd.read_excel(file_path, sheet_name=desired_sheet[0], skiprows=start_row-1, usecols=f"A:{end_col}", nrows=end_row-start_row+1)
                print("DATAFRAME1112223334455566--->",df)
                # Find the row index where 'Total' is present in any cell
                total_row_index = df.apply(lambda row: row.astype(str).str.contains('UNITS|UNTIS', case=False).any(), axis=1).idxmax()

                # Exclude the row with 'Total' if found
                if total_row_index is not None:
                    df = df.loc[:total_row_index - 4]

                print(f"DataFrame for <{desired_sheet[0]}> :>>>>>>", df)
                return df

            elif "INVOICE" in desired_sheet[0]:
                # Specify the range of cells in the table
                start_row = 13  # Replace with the actual starting row    of your table
                end_row = 23    # Replace with the actual ending   row    of your table
                start_col = 'A' # Replace with the actual starting column of your table
                end_col = 'N'   # Replace with the actual ending   column of your table

                # Read the data from the desired sheet and specified range
                df = pd.read_excel(file_path, sheet_name=desired_sheet[0], skiprows=start_row-1, usecols=f"{start_col}:{end_col}", nrows=end_row-start_row+1)

                # Find the row index where 'Total' is present in any cell
                total_row_index = df.apply(lambda row: row.astype(str).str.contains('Total', case=False).any(), axis=1).idxmax()

                # Exclude the row with 'Total' if found
                if total_row_index is not None:
                    df = df.loc[:total_row_index - 1]

                # df = df.fillna('')
                # Display the DataFrame or perform further operations as needed
                print(f"DataFrame for <{desired_sheet[0]}> :>>>>>>",df)

                return df
        except:
            # Specify the range of cells in the table
            df = pd.read_excel(file_path, sheet_name=desired_sheet[0])
            start_row = 28  # Replace with the actual starting row of your table
            end_row = 34   # Replace with the actual ending row of your table
            start_col = 'A' # Replace with the actual starting column of your table
            end_col = chr(ord('A') + df.shape[1] - 1)
            df = pd.read_excel(file_path, sheet_name=desired_sheet[0], skiprows=start_row-1, usecols=f"A:{end_col}", nrows=end_row-start_row+1)
            print("DATAFRAME no medoator -->",df)
            # Find the row index where 'Total' is present in any cell
            total_row_index = df.apply(lambda row: row.astype(str).str.contains('Total', case=False).any(), axis=1).idxmax()

            # Exclude the row with 'Total' if found
            if total_row_index is not None:
                df = df.loc[:total_row_index - 1]

            # df = df.fillna('')
            # Display the DataFrame or perform further operations as needed
            print(f"DataFrame no medoator for <{desired_sheet[0]}> :>>>>>>",df)
            return df
    
